import pandas as pd
import numpy as np
import os
import sys
import time
import win32com.client
from openpyxl import load_workbook
from sklearn.exceptions import ConvergenceWarning
from sklearn.preprocessing import RobustScaler
from sklearn.linear_model import LogisticRegression, Lasso, LinearRegression
from sklearn import linear_model
from sklearn import svm
from sklearn import tree
from sklearn.neighbors import KNeighborsClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import mean_absolute_error,mean_squared_error, accuracy_score,balanced_accuracy_score, confusion_matrix, ConfusionMatrixDisplay
import json
from sklearn.model_selection import train_test_split,GridSearchCV, KFold,LeaveOneOut
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_classif
from sklearn.feature_selection import SequentialFeatureSelector
from sklearn.model_selection import cross_val_score
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.multioutput import MultiOutputClassifier
from sklearn.metrics import hamming_loss
from sklearn.feature_selection import SelectFromModel
from sklearn.preprocessing import StandardScaler
import pickle
import glob
import matplotlib.pyplot as plt


def leave_one_participant_out_split(df, participant, features, validation):
    """ One participant is held out for validation, rest are trained on """
    # Train: all participants but 1
    
    train_data = df.loc[df.id != participant]
    train = train_data[features]
    y_train = np.ravel(train_data[validation])
    
    # Test: the held out participant
    test_data = df.loc[df.id == participant]
    test = test_data[features]
    y_test = np.ravel(test_data[validation])

    return train, y_train, test, y_test


def get_clf(clf_key):
    """ All of the ML models are here
    You can also fine tune parameters for each model here"""
    # internal dependency with clf_keys list
    if clf_key == 'tree':
        clf = tree.DecisionTreeClassifier(max_depth=5)
    elif clf_key == 'svc':
        clf = svm.SVC(kernel='rbf')
    elif clf_key == 'logreg':
        clf = LogisticRegression(solver='saga')
    elif clf_key == 'knn':
        clf = KNeighborsClassifier(n_neighbors=9)
    elif clf_key == 'mlp':
        clf = MLPClassifier(solver='adam',max_iter=10000)
    elif clf_key == 'nb':
        clf = GaussianNB()
    elif clf_key == 'rf':
        clf = RandomForestClassifier()
    elif clf_key == 'lr':
        clf = LinearRegression()
    elif clf_key == 'lar':
        clf = linear_model.Lasso(alpha=0.1)
    else:
        print('Not a valid clf key')
        sys.exit()

    return clf

def save_accuracy(file_name, df_scores, clf_keys):
    """ Saves all the accuracies of the ML model on each participant
    to an excel file """
    first_file = 'please_delete.xlsx'
    
    df_scores.to_excel(first_file)
    
    xls = pd.ExcelFile(first_file)
    
    # Load the first sheet into a DataFrame
    df = xls.parse(xls.sheet_names[0])
    
    df.insert(5, '', pd.NA)
    df.insert(6, 'Model', pd.NA)
    df.insert(7, 'Average', pd.NA)
    df.insert(8, 'Min', pd.NA)  
    df.insert(9, 'Max', pd.NA) 
    
    # Save the DataFrame back to the Excel file temporarily
    with pd.ExcelWriter(first_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name=xls.sheet_names[0])

    # Load the workbook and the specific sheet
    workbook = load_workbook(first_file)
    sheet = workbook[xls.sheet_names[0]]
    
    # Model Names
    for row in range(2, len(clf_keys)+2):
        sheet[f'G{row}'] = clf_keys[row-2]
        
    # Averages
    for row in range(2, len(clf_keys)+2):
        sheet[f'H{row}'] = f'=AVERAGEIF(C:C,G{row},D:D)'    
    
    # Min Value
    min_val = 1.0
    for ml in range(2, len(clf_keys)+2):
        for row in range (2 + (ml-2), sheet.max_row + 1, len(clf_keys)):
            cell_value = sheet[f'D{row}'].value
            if cell_value is not None and float(cell_value) < min_val:
                min_val = float(cell_value)
        sheet[f'I{ml}'] = min_val
        min_val = 1.0
    
    # Max Value
    max_val = 0.0
    for ml in range(2, len(clf_keys)+2):
        for row in range (2 + (ml-2), sheet.max_row + 1, len(clf_keys)):
            cell_value = sheet[f'D{row}'].value
            if cell_value is not None and float(cell_value) > max_val:
                max_val = float(cell_value)
        sheet[f'J{ml}'] = max_val
        max_val = 0.0           
    
    workbook.save(first_file)
    refresh(first_file, file_name)
    print(f'saved to {file_name}')


def refresh(file_path, sheet_name):
    """ Refreshes excel file to allow for equations to work correctly with pandas """
    # Start an instance of Excel
    xlapp = win32com.client.DispatchEx("Excel.Application")

    # Open the workbook in said instance of Excel
    wb = xlapp.workbooks.open(f'{os.getcwd()}\{file_path}')

    # Optional, e.g. if you want to debug
    # xlapp.Visible = True
    sheet = f'{os.getcwd()}\{sheet_name}'
    # Refresh all data connections.
    wb.RefreshAll()
    wb.SaveAs(sheet)
    # Quit
    xlapp.Quit()

def makeDict(conditions):
    temp = {}
    i=1
    for x in conditions:
        temp[x] = i
        i+=1
    return temp


def testConditions(file, conditions):
    reduced_file_name = file[file.rfind('\\')+1:].split('.')[0]
    print(f'Starting Test For {conditions[0]} and {conditions[1]}, for {reduced_file_name}.csv')
    """ Generates and tests an ensemble classifier for laughter recognition """
    # contains all the data we need for classification
    master_df = pd.read_csv(file)
   
    master_df = master_df[master_df['condition'].str.contains('|'.join(conditions))]
    
    master_df['id'] =  master_df['id'].str.slice(1).astype(int)

    master_df['condition'] = master_df['condition'].map(makeDict(conditions))

   # print(master_df)
 
    # # training features
    # Features List (formatted_features is all features)
    columns = np.array(master_df.columns.values)
    if 'Unnamed: 0' in columns:
        formatted_features = columns[3:]
    else: formatted_features = columns[2:]
    #intensity_features = ['Yaw','Pitch','Roll','AvgVelocity','MaxVelocity','AvgAccel','MaxAccel','AvgArea','COGZ','PathLengthDifference','RangeX','RangeY','RangeZ',
    #                      'TimeElapsed'
    #                      ]
    
    
    # ground truth validation column
    validation = ['condition']
 
    # MAKE IT 2 CLASS -1/0 vs 1
    # master_df[validation] = master_df[validation].replace(-1, 0)

    clf_keys = ['logreg', 'svc', 'knn','mlp','nb']
    #clf_keys = ['lar','lr','logreg']


    y = np.ravel(master_df[validation])
    X = master_df[formatted_features]
    #print(X)

    transformer = RobustScaler().fit(X)
    new_x = transformer.transform(X)


    test_frame = pd.DataFrame(new_x,columns=formatted_features)

    master_df.drop(labels=formatted_features, axis="columns", inplace=True)
    master_df.reset_index(drop=True, inplace=True)

    master_df[formatted_features] = test_frame[formatted_features]

    #calling the model with the best parameter

    df_scores = pd.DataFrame()
    scores = []

    #Create For Confusion Matrix
    cm_total = []
    for i in range(len(conditions)):
        cm_total.append(np.zeros(len(conditions)))
    
    for i in range(2,31):
        #AB 6 and 21
        if i == 4 or i==6 or i ==21:
            pass
        else:
            #print(i)
            X_train,y_train,X_test,y_test = leave_one_participant_out_split(master_df, i, formatted_features, validation)
            #print('pls', X_test)

            for clf_key in clf_keys:
                clf = get_clf(clf_key)
                #X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.1, random_state=1066)

                clf.fit(X_train, y_train)                
                #print("Model and scaler have been saved")
                results=clf.predict(X_test)
                error = mean_absolute_error(y_test,results)
                accuracy = accuracy_score(y_test,results)
                
                cm = confusion_matrix(y_test, results,labels=range(1,len(conditions)+1))

                cm_total = np.array(cm) + np.array(cm_total)
                rows = {"Test Participant": i,"Classifier":clf_key,"Accuracy": accuracy,"Mean Error": error}
                df_scores = pd.concat([df_scores, pd.DataFrame({
                    "Test Participant": [i],
                    "Classifier": [clf_key],
                    "Accuracy": [accuracy],
                    "Mean Error": [error]
                })], ignore_index=True)
    
   
    #newpath = r'C:\Program Files\arbitrary' 
    new_pth = os.getcwd()  
    new_pth = os.path.join(new_pth,'OUTPUTS')
    new_pth = os.path.join(new_pth,reduced_file_name)
    
    if not os.path.exists(new_pth):
        os.makedirs(new_pth)
    print(df_scores)
    cm_display = ConfusionMatrixDisplay(cm_total, display_labels=conditions).plot()
    title = f'CM_{conditions[0],conditions[1]}'
    cm_display.ax_.set(xlabel='Predicted', ylabel='True',title=title)
    plts_pth = os.path.join(new_pth,'plots')
    if not os.path.exists(plts_pth):
        os.makedirs(plts_pth)
    plt.savefig(f'{plts_pth}\{title}.png')
    #plt.show()
    files_pth = os.path.join(new_pth,'files')
    if not os.path.exists(files_pth):
        os.makedirs(files_pth)
    

    
    save_accuracy(f'OUTPUTS\{reduced_file_name}\\files\{conditions[0]}_{conditions[1]}.xlsx', df_scores, clf_keys)

    print(f"Completed run of {conditions[0]}_{conditions[1]}, from {reduced_file_name}.csv")
    return formatted_features, f'{reduced_file_name}.csv', df_scores

def check_string_positions(small_list, large_list):
    result = []
    for string in small_list:
        if string in large_list:
            result.append('1')
        else:
            result.append('0')
    return result


if __name__ == "__main__":
    ref_file = 'C:\\Users\\nnamd\Documents\\GitHub\\kinova_share\\data_analysis\\pull_from\\ConfidenceFilter\\reduced.csv'
    #Get Column Data
    feature_data = pd.read_csv(ref_file)
    columns = np.array(feature_data.columns.values)
    if 'Unnamed: 0' in columns:
        all_features = columns[3:]
    else: all_features = columns[2:]


    path = "C:/Users/nnamd/Documents/GitHub/kinova_share/data_analysis/pull_from/ConfidenceFilter/*.csv"
    files = []
    compare = []
    for_conven = pd.DataFrame()

    for file in glob.glob(path):
        #print(file)
        features, file_name, df_scores = testConditions(file, ['A','B'])
        

        features, file_name, df_scores = testConditions(file, ['C','D'])
        features, file_name, df_scores = testConditions(file, ['E','F'])
        features, file_name, df_scores = testConditions(file, ['G','H'])
        
        
        
        bools = check_string_positions(features, all_features)
        compare.append(bools)
        files.append(file_name)
    for_conven['filename'] = np.array(files)
    for_conven[all_features] = np.array(compare)
    for_conven.to_excel('model_accuracy_results.xlx')

    # """ Generates and tests an ensemble classifier for laughter recognition """
    # # contains all the data we need for classification
    # conditions = ['A','B']
    # master_df = pd.read_csv('C:\\Users\\nnamd\Documents\\GitHub\\kinova_share\\data_analysis\\pull_from\\ConfidenceFilter\\robot_data_removal.csv')
    # master_df = master_df[master_df['condition'].str.contains('|'.join(conditions))]
    
    # master_df['id'] =  master_df['id'].str.slice(1).astype(int)
    # print('main',master_df)
    # master_df['condition'] = master_df['condition'].map(makeDict(conditions))
    # print('post',master_df)
    
 
    # # # training features
    # # Features List (formatted_features is all features)
    # formatted_features = ['Yaw','Pitch','Roll','AvgVelocity','MaxVelocity','AvgAccel','MaxAccel','AvgArea','COGZ','PathLengthDifference','RangeX','RangeY','RangeZ',
    #                       'TimeElapsed'
    #                       ]
    # intensity_features = ['Yaw','Pitch','Roll','AvgVelocity','MaxVelocity','AvgAccel','MaxAccel','AvgArea','COGZ','PathLengthDifference','RangeX','RangeY','RangeZ',
    #                       'TimeElapsed'
    #                       ]
    
    
    # # ground truth validation column
    # validation = ['condition']
 
    # # MAKE IT 2 CLASS -1/0 vs 1
    # # master_df[validation] = master_df[validation].replace(-1, 0)

    # clf_keys = ['logreg', 'svc', 'knn']
    # #clf_keys = ['lar','lr','logreg']


    # y = np.ravel(master_df[validation])
    # X = master_df[formatted_features]
    # print('post-hoc',X)

    # transformer = RobustScaler().fit(X)
    # new_x = transformer.transform(X)
    # print('new_x',new_x)

    # test_frame = pd.DataFrame(new_x,columns=formatted_features)

    # master_df.drop(labels=formatted_features, axis="columns", inplace=True)
    # master_df.reset_index(drop=True, inplace=True)

    # master_df[formatted_features] = test_frame[formatted_features]

    # #calling the model with the best parameter

    # df_scores = pd.DataFrame()
    # scores = []

    # #Create For Confusion Matrix
    # cm_total = []
    # for i in range(len(conditions)):
    #     cm_total.append(np.zeros(len(conditions)))
    
    # for i in range(2,31):
    #     #AB 6 and 21
    #     if i == 4 or i==6 or i ==21:
    #         pass
    #     else:
    #         print(i)
    #         X_train,y_train,X_test,y_test = leave_one_participant_out_split(master_df, i, formatted_features, validation)
    #         print(X_train)

    #         for clf_key in clf_keys:
    #             clf = get_clf(clf_key)
    #             #X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=.1, random_state=1066)

    #             clf.fit(X_train, y_train)                
    #             print("Model and scaler have been saved")
    #             print(i, X_test)
    #             results=clf.predict(X_test)
    #             error = mean_absolute_error(y_test,results)
    #             accuracy = accuracy_score(y_test,results)
                
    #             cm = confusion_matrix(y_test, results,labels=range(1,len(conditions)+1),normalize='true')

    #             cm_total = np.array(cm) + np.array(cm_total)
    #             rows = {"Test Participant": i,"Classifier":clf_key,"Accuracy": accuracy,"Mean Error": error}
    #             df_scores = pd.concat([df_scores, pd.DataFrame({
    #                 "Test Participant": [i],
    #                 "Classifier": [clf_key],
    #                 "Accuracy": [accuracy],
    #                 "Mean Error": [error]
    #             })], ignore_index=True)
    

    # print(df_scores)
    # cm_display = ConfusionMatrixDisplay(cm_total, display_labels=conditions).plot()
    # title = f'Tree Joke Success - Mode'
    # cm_display.ax_.set(xlabel='Predicted', ylabel='True',title=title)
    # #plt.savefig(f'.\Accuracy-CMs\{title}.png')
    # plt.show()
    
    # save_accuracy(f'{validation[0].lower()}_formatted_1000.xlsx')
    